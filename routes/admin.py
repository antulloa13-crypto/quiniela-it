"""
Rutas del panel de administración.
Requiere rol 'admin'. Gestiona participantes, partidos, resultados y ranking.
"""
from functools import wraps
from datetime import datetime
import io

from flask import Blueprint, render_template, redirect, url_for, request, flash, abort, send_file
from flask_login import login_required, current_user

from models import User, Partido, Pronostico, Resultado, Puntuacion
from extensions import db
from utils import calcular_puntuaciones, get_ranking

admin_bp = Blueprint('admin', __name__)


def admin_required(f):
    """Decorador que exige rol admin. Devuelve 403 si no cumple."""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.es_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ─── Dashboard ────────────────────────────────────────────────────────────────

@admin_bp.route('/')
@admin_bp.route('/dashboard')
@admin_required
def dashboard():
    total_participantes = User.query.filter_by(rol='participante', activo=True).count()
    total_partidos = Partido.query.count()
    partidos_pendientes = Partido.query.filter_by(estado='pendiente').count()
    partidos_en_juego = Partido.query.filter_by(estado='en_juego').count()
    partidos_finalizados = Partido.query.filter_by(estado='finalizado').count()

    ranking = get_ranking()[:10]

    ultimos_resultados = (
        Resultado.query
        .join(Partido, Resultado.partido_id == Partido.id)
        .order_by(Resultado.fecha_registro.desc())
        .limit(5)
        .all()
    )

    return render_template(
        'admin/dashboard.html',
        total_participantes=total_participantes,
        total_partidos=total_partidos,
        partidos_pendientes=partidos_pendientes,
        partidos_en_juego=partidos_en_juego,
        partidos_finalizados=partidos_finalizados,
        ranking=ranking,
        ultimos_resultados=ultimos_resultados,
    )


# ─── Participantes ────────────────────────────────────────────────────────────

@admin_bp.route('/participantes')
@admin_required
def participantes():
    lista = User.query.filter_by(rol='participante').order_by(User.apellido).all()
    return render_template('admin/participantes.html', participantes=lista)


@admin_bp.route('/participantes/nuevo', methods=['GET', 'POST'])
@admin_required
def nuevo_participante():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        apellido = request.form.get('apellido', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        errors = _validar_participante(nombre, apellido, username, password)

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/participante_form.html',
                                   titulo='Nuevo Participante',
                                   user=None,
                                   form_data=request.form)

        user = User(nombre=nombre, apellido=apellido, username=username, rol='participante')
        user.set_password(password)
        db.session.add(user)
        db.session.flush()

        puntuacion = Puntuacion(usuario_id=user.id)
        db.session.add(puntuacion)
        db.session.commit()

        flash(f'Participante <strong>{username}</strong> creado correctamente.', 'success')
        return redirect(url_for('admin.participantes'))

    return render_template('admin/participante_form.html',
                           titulo='Nuevo Participante',
                           user=None,
                           form_data={})


@admin_bp.route('/participantes/<int:user_id>/editar', methods=['GET', 'POST'])
@admin_required
def editar_participante(user_id):
    user = User.query.get_or_404(user_id)
    if user.es_admin:
        flash('No se puede editar la cuenta de administrador desde aquí.', 'warning')
        return redirect(url_for('admin.participantes'))

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        apellido = request.form.get('apellido', '').strip()
        password = request.form.get('password', '').strip()
        activo = request.form.get('activo') == 'on'

        errors = []
        if not nombre:
            errors.append('El nombre es requerido.')
        if not apellido:
            errors.append('El apellido es requerido.')
        if password and len(password) < 6:
            errors.append('La nueva contraseña debe tener al menos 6 caracteres.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/participante_form.html',
                                   titulo='Editar Participante',
                                   user=user,
                                   form_data=request.form)

        user.nombre = nombre
        user.apellido = apellido
        user.activo = activo
        if password:
            user.set_password(password)

        db.session.commit()
        flash(f'Participante <strong>{user.username}</strong> actualizado.', 'success')
        return redirect(url_for('admin.participantes'))

    form_data = {
        'nombre': user.nombre,
        'apellido': user.apellido,
        'activo': user.activo,
    }
    return render_template('admin/participante_form.html',
                           titulo='Editar Participante',
                           user=user,
                           form_data=form_data)


@admin_bp.route('/participantes/<int:user_id>/eliminar', methods=['POST'])
@admin_required
def eliminar_participante(user_id):
    user = User.query.get_or_404(user_id)
    if user.es_admin:
        flash('No se puede eliminar la cuenta de administrador.', 'danger')
        return redirect(url_for('admin.participantes'))

    username = user.username
    db.session.delete(user)
    db.session.commit()
    # Recalcular tras eliminar participante
    calcular_puntuaciones()
    flash(f'Participante <strong>{username}</strong> eliminado.', 'success')
    return redirect(url_for('admin.participantes'))


# ─── Partidos ─────────────────────────────────────────────────────────────────

@admin_bp.route('/partidos')
@admin_required
def partidos():
    lista = Partido.query.order_by(Partido.fecha_hora.desc()).all()
    return render_template('admin/partidos.html', partidos=lista)


@admin_bp.route('/partidos/nuevo', methods=['GET', 'POST'])
@admin_required
def nuevo_partido():
    if request.method == 'POST':
        partido, errors = _partido_desde_form(request.form)
        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/partido_form.html',
                                   titulo='Nuevo Partido',
                                   partido=None,
                                   form_data=request.form)

        db.session.add(partido)
        db.session.commit()
        flash(f'Partido <strong>{partido.equipo_a} vs {partido.equipo_b}</strong> creado.', 'success')
        return redirect(url_for('admin.partidos'))

    return render_template('admin/partido_form.html',
                           titulo='Nuevo Partido',
                           partido=None,
                           form_data={'estado': 'pendiente'})


@admin_bp.route('/partidos/<int:partido_id>/editar', methods=['GET', 'POST'])
@admin_required
def editar_partido(partido_id):
    partido = Partido.query.get_or_404(partido_id)

    if request.method == 'POST':
        nuevo, errors = _partido_desde_form(request.form)
        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('admin/partido_form.html',
                                   titulo='Editar Partido',
                                   partido=partido,
                                   form_data=request.form)

        partido.equipo_a = nuevo.equipo_a
        partido.equipo_b = nuevo.equipo_b
        partido.fecha_hora = nuevo.fecha_hora
        partido.estado = nuevo.estado
        db.session.commit()
        flash('Partido actualizado correctamente.', 'success')
        return redirect(url_for('admin.partidos'))

    form_data = {
        'equipo_a': partido.equipo_a,
        'equipo_b': partido.equipo_b,
        'fecha_hora': partido.fecha_hora.strftime('%Y-%m-%dT%H:%M'),
        'estado': partido.estado,
    }
    return render_template('admin/partido_form.html',
                           titulo='Editar Partido',
                           partido=partido,
                           form_data=form_data)


@admin_bp.route('/partidos/<int:partido_id>/eliminar', methods=['POST'])
@admin_required
def eliminar_partido(partido_id):
    partido = Partido.query.get_or_404(partido_id)
    nombre = f"{partido.equipo_a} vs {partido.equipo_b}"
    db.session.delete(partido)
    db.session.commit()
    calcular_puntuaciones()
    flash(f'Partido <strong>{nombre}</strong> eliminado.', 'success')
    return redirect(url_for('admin.partidos'))


# ─── Resultados ───────────────────────────────────────────────────────────────

@admin_bp.route('/resultados')
@admin_required
def resultados():
    todos = Partido.query.order_by(Partido.fecha_hora.asc()).all()
    sin_resultado = [p for p in todos if p.resultado is None]
    con_resultado = [p for p in todos if p.resultado is not None]
    return render_template(
        'admin/resultados.html',
        partidos=todos,
        sin_resultado=sin_resultado,
        con_resultado=con_resultado,
    )


@admin_bp.route('/resultados/<int:partido_id>', methods=['POST'])
@admin_required
def guardar_resultado(partido_id):
    partido = Partido.query.get_or_404(partido_id)
    resultado_val = request.form.get('resultado', '').strip()

    if resultado_val not in ('equipo_a', 'empate', 'equipo_b'):
        flash('Resultado inválido.', 'danger')
        return redirect(url_for('admin.resultados'))

    # Actualizar o crear resultado
    resultado = Resultado.query.filter_by(partido_id=partido_id).first()
    if resultado is None:
        resultado = Resultado(partido_id=partido_id, resultado=resultado_val)
        db.session.add(resultado)
    else:
        resultado.resultado = resultado_val
        resultado.fecha_registro = datetime.now()

    partido.estado = 'finalizado'
    db.session.commit()

    # Recalcular todas las puntuaciones
    calcular_puntuaciones()

    flash(
        f'Resultado registrado: <strong>{partido.equipo_a} vs {partido.equipo_b}</strong>.',
        'success'
    )
    return redirect(url_for('admin.resultados'))


# ─── Ranking ──────────────────────────────────────────────────────────────────

@admin_bp.route('/ranking')
@admin_required
def ranking():
    data = get_ranking()
    return render_template('ranking.html', ranking=data, is_admin=True)


# ─── Importar pronósticos desde Excel ─────────────────────────────────────────

@admin_bp.route('/pronosticos/plantilla')
@admin_required
def descargar_plantilla():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Pronósticos"

    # ── estilos ──
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    ref_fill    = PatternFill("solid", fgColor="D9E8F5")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    ref_font    = Font(color="555555", italic=True, size=10)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["usuario", "partido_id", "equipo_a", "vs", "equipo_b", "pronostico"]
    col_widths = [18, 12, 22, 5, 22, 18]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    # ── datos ──
    usuarios = User.query.filter_by(rol='participante', activo=True).order_by(User.apellido).all()
    partidos = Partido.query.order_by(Partido.fecha_hora).all()

    row_num = 2
    for user in usuarios:
        for partido in partidos:
            row_data = [
                user.username,
                partido.id,
                partido.equipo_a,
                "vs",
                partido.equipo_b,
                "",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = border
                cell.alignment = center if col_idx in (2, 4) else left
                # columnas de referencia (partido_id, equipo_a, vs, equipo_b) → gris claro
                if col_idx in (2, 3, 4, 5):
                    cell.fill = ref_fill
                    cell.font = ref_font
                else:
                    cell.fill = white_fill
            row_num += 1

    # ── validación desplegable para columna pronostico (F) ──
    if row_num > 2:
        dv = DataValidation(
            type="list",
            formula1='"equipo_a,empate,equipo_b"',
            allow_blank=True,
            showErrorMessage=True,
            error="Use: equipo_a, empate o equipo_b",
            errorTitle="Valor inválido",
        )
        ws.add_data_validation(dv)
        dv.sqref = f"F2:F{row_num - 1}"

    ws.freeze_panes = "A2"

    # ── hoja de instrucciones ──
    wi = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("INSTRUCCIONES DE USO", True),
        ("", False),
        ("1. Completa SOLO la columna 'pronostico' (columna F).", False),
        ("2. Valores permitidos:", False),
        ("   • equipo_a  → gana el equipo de la columna 'equipo_a'", False),
        ("   • equipo_b  → gana el equipo de la columna 'equipo_b'", False),
        ("   • empate    → el partido termina en empate", False),
        ("", False),
        ("3. No modifiques las columnas grises (son de referencia).", False),
        ("4. Puedes dejar filas en blanco si no deseas importar ese pronóstico.", False),
        ("5. Si el usuario ya tenía un pronóstico para ese partido, se actualizará.", False),
    ]
    wi.column_dimensions["A"].width = 65
    for i, (text, bold) in enumerate(instrucciones, start=1):
        cell = wi.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=11 if bold else 10,
                         color="1A3A5C" if bold else "333333")
        cell.alignment = left

    # ── enviar archivo ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="plantilla_pronosticos.xlsx",
    )


@admin_bp.route('/pronosticos/importar', methods=['GET', 'POST'])
@admin_required
def importar_pronosticos():
    if request.method == 'POST':
        archivo = request.files.get('archivo')
        if not archivo or not archivo.filename.lower().endswith('.xlsx'):
            flash('Selecciona un archivo .xlsx válido.', 'danger')
            return redirect(url_for('admin.importar_pronosticos'))

        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(archivo.read()), data_only=True)
            ws = wb.active
        except Exception:
            flash('No se pudo leer el archivo Excel. Asegúrate de que sea un .xlsx válido.', 'danger')
            return redirect(url_for('admin.importar_pronosticos'))

        errores = []
        nuevos = 0
        actualizados = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            username     = str(row[0]).strip() if row[0] is not None else ''
            partido_id_v = row[1]
            pronostico_v = str(row[5]).strip().lower() if row[5] is not None else ''

            if not username or not partido_id_v or not pronostico_v:
                continue  # fila vacía o incompleta → ignorar

            if pronostico_v not in ('equipo_a', 'empate', 'equipo_b'):
                errores.append(
                    f"Fila {row_num}: pronóstico inválido '{pronostico_v}'. "
                    "Use: equipo_a, empate o equipo_b."
                )
                continue

            user = User.query.filter_by(username=username, rol='participante').first()
            if not user:
                errores.append(f"Fila {row_num}: usuario '{username}' no encontrado.")
                continue

            try:
                partido_id_int = int(partido_id_v)
            except (ValueError, TypeError):
                errores.append(f"Fila {row_num}: partido_id inválido '{partido_id_v}'.")
                continue

            partido = Partido.query.get(partido_id_int)
            if not partido:
                errores.append(f"Fila {row_num}: partido con ID {partido_id_int} no existe.")
                continue

            existing = Pronostico.query.filter_by(
                usuario_id=user.id, partido_id=partido.id
            ).first()

            if existing:
                existing.pronostico = pronostico_v
                existing.fecha_registro = datetime.now()
                actualizados += 1
            else:
                db.session.add(Pronostico(
                    usuario_id=user.id,
                    partido_id=partido.id,
                    pronostico=pronostico_v,
                ))
                nuevos += 1

        if nuevos or actualizados:
            db.session.commit()
            calcular_puntuaciones()
            flash(
                f'Importación completada: <strong>{nuevos}</strong> nuevos pronósticos, '
                f'<strong>{actualizados}</strong> actualizados.',
                'success',
            )
        else:
            if not errores:
                flash('El archivo no contenía datos para importar.', 'warning')

        for e in errores:
            flash(e, 'warning')

        return redirect(url_for('admin.importar_pronosticos'))

    return render_template('admin/importar_pronosticos.html')


# ─── Helpers privados ─────────────────────────────────────────────────────────

def _validar_participante(nombre, apellido, username, password, user_id=None):
    errors = []
    if not nombre:
        errors.append('El nombre es requerido.')
    if not apellido:
        errors.append('El apellido es requerido.')
    if not username:
        errors.append('El nombre de usuario es requerido.')
    elif not all(c.isalnum() or c == '_' for c in username):
        errors.append('El usuario solo puede contener letras, números y guiones bajos (_).')
    else:
        existing = User.query.filter_by(username=username).first()
        if existing and (user_id is None or existing.id != user_id):
            errors.append(f'El usuario "{username}" ya está en uso.')
    if not password:
        errors.append('La contraseña es requerida.')
    elif len(password) < 6:
        errors.append('La contraseña debe tener al menos 6 caracteres.')
    return errors


def _partido_desde_form(form):
    """Construye un Partido desde el form y retorna (partido, lista_errores)."""
    equipo_a = form.get('equipo_a', '').strip()
    equipo_b = form.get('equipo_b', '').strip()
    fecha_hora_str = form.get('fecha_hora', '').strip()
    estado = form.get('estado', 'pendiente').strip()

    errors = []
    fecha_hora = None

    if not equipo_a:
        errors.append('El nombre del Equipo A es requerido.')
    if not equipo_b:
        errors.append('El nombre del Equipo B es requerido.')
    if equipo_a and equipo_b and equipo_a.lower() == equipo_b.lower():
        errors.append('Los equipos no pueden tener el mismo nombre.')
    if not fecha_hora_str:
        errors.append('La fecha y hora son requeridas.')
    else:
        try:
            fecha_hora = datetime.strptime(fecha_hora_str, '%Y-%m-%dT%H:%M')
        except ValueError:
            errors.append('Formato de fecha/hora inválido.')
    if estado not in ('pendiente', 'en_juego', 'finalizado'):
        errors.append('Estado inválido.')

    if errors:
        return None, errors

    partido = Partido(
        equipo_a=equipo_a,
        equipo_b=equipo_b,
        fecha_hora=fecha_hora,
        estado=estado,
    )
    return partido, []
